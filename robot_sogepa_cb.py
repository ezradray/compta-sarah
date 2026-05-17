import pdfplumber
import re
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

DATE_RE  = re.compile(r'^(\d{2}/\d{2}/\d{4})\s+(.+?)\s+(Avoir\s+)?([\d\.]+,\d{2})\s*$')
TOTAL_RE = re.compile(r'TOTAL PRELEVE.*?:\s*([\d\s\.]+,\d{2})\s*euros', re.IGNORECASE)
TITU_RE  = re.compile(r'TITULAIRE\s*:\s*(.+)', re.IGNORECASE)
CARTE_RE = re.compile(r'CARTE Business N°\s*([\dX\s]+?)\s+Au', re.IGNORECASE)

def _parse_pdf(pdf_io):
    """Lit le PDF et retourne dict {num_carte: {titulaire, lignes, total}}"""
    cartes = {}
    num_courant = None
    pdf_io.seek(0)
    with pdfplumber.open(pdf_io) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text: continue
            for line in text.split('\n'):
                line = line.strip()
                m_carte = CARTE_RE.search(line)
                if m_carte:
                    num = m_carte.group(1).strip()
                    num_courant = num
                    if num not in cartes:
                        cartes[num] = {'titulaire':'','lignes':[],'total':0.0}
                    continue
                m_tit = TITU_RE.match(line)
                if m_tit and num_courant:
                    if not cartes[num_courant]['titulaire']:
                        cartes[num_courant]['titulaire'] = m_tit.group(1).strip()
                    continue
                m_tot = TOTAL_RE.search(line)
                if m_tot and num_courant:
                    tot_str = m_tot.group(1).replace(' ','').replace('.','').replace(',','.')
                    cartes[num_courant]['total'] = float(tot_str)
                    num_courant = None
                    continue
                if num_courant:
                    m = DATE_RE.match(line)
                    if m:
                        try:
                            montant = float(m.group(4).replace('.','').replace(',','.'))
                            avoir = bool(m.group(3))
                            if avoir: montant = -montant
                            cartes[num_courant]['lignes'].append({
                                'date'   : m.group(1),
                                'libelle': m.group(2).strip(),
                                'montant': montant
                            })
                        except: pass
    return cartes

def run_sogepa_cb(pdf_io, pivot_io, mois, annee=2026):
    """
    Entrée  : PDF relevé cartes SOGEPA, Excel pivot, mois, annee
    Sortie  : (buf_excel, stats) ou (None, message_erreur)
    """
    # ── PIVOT — fournisseurs 401xxx ───────────────────────────
    pivot_io.seek(0)
    df_p = __import__('pandas').read_excel(pivot_io, header=None)
    cols = ['COMPTE','LIBELLE'] + [f'C{i}' for i in range(len(df_p.columns)-2)]
    df_p.columns = cols
    df_p = df_p[['COMPTE','LIBELLE']].dropna(subset=['LIBELLE'])
    df_p['COMPTE']  = df_p['COMPTE'].astype(str).str.strip()
    df_p['LIBELLE'] = df_p['LIBELLE'].astype(str).str.strip()
    fournisseurs = df_p[df_p['COMPTE'].str.startswith('401')].copy()

    def _get_compte(libelle):
        """46700000 par défaut, compte pivot si fournisseur reconnu"""
        lib_up = libelle.upper()
        for _, r in fournisseurs.iterrows():
            mots = [m for m in re.split(r'\W+', r['LIBELLE'].upper()) if len(m) > 3]
            if any(m in lib_up for m in mots):
                return r['COMPTE']
        return '46700000'

    # ── LECTURE PDF ───────────────────────────────────────────
    cartes = _parse_pdf(pdf_io)
    if not cartes:
        return None, "Aucune carte trouvée dans le PDF"

    # Vérification totaux
    alertes = []
    for num, c in cartes.items():
        somme = sum(l['montant'] for l in c['lignes'])
        if abs(somme - c['total']) > 0.05:
            alertes.append(f"⚠️ Carte {num} ({c['titulaire']}) : somme={somme:.2f} ≠ total={c['total']:.2f}")

    # ── GÉNÉRATION EXCEL ──────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = 'Feuil1'

    headers = ['DATE','CODE BANQUE','COMPTE','PIECE','LIBELLE','DEBIT','CREDIT']
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = PatternFill('solid', fgColor='1E3A5F')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')

    f1 = PatternFill('solid', fgColor='EEF2FF')
    f2 = PatternFill('solid', fgColor='FFFFFF')
    row_idx = 2

    for num, carte in cartes.items():
        titulaire = carte['titulaire']
        total     = carte['total']
        date_prel = f"30/{mois:02d}/{annee}"  # date prélèvement

        for ligne in carte['lignes']:
            d, m, y = ligne['date'].split('/')
            date_iso = f"{y}-{m}-{d}"
            # Libellé = Commerce + date
            libelle  = f"{ligne['libelle']} {ligne['date']}"[:50]
            compte   = _get_compte(ligne['libelle'])
            montant  = ligne['montant']
            fill     = f1 if (row_idx-2) % 2 == 0 else f2

            if montant >= 0:
                ws.append([date_iso,'BQ0', compte,    'CB', libelle, round(montant,2), 0])
            else:
                ws.append([date_iso,'BQ0', compte,    'CB', libelle+' (Avoir)', 0, round(abs(montant),2)])
            for cell in ws[row_idx]: cell.fill = fill
            row_idx += 1

        # Contrepartie banque
        d2, m2, y2 = date_prel.split('/')
        date_iso2 = f"{y2}-{m2}-{d2}"
        lib_total = f"RELEVE CARTE {num} {titulaire} {date_prel}"[:50]
        fill = f1 if (row_idx-2) % 2 == 0 else f2
        ws.append([date_iso2,'BQ0','51200000','CB', lib_total, 0, round(total,2)])
        for cell in ws[row_idx]: cell.fill = fill
        row_idx += 1

    # Colonnes largeur
    for col, w in [('A',14),('B',10),('C',12),('D',8),('E',55),('F',14),('G',14)]:
        ws.column_dimensions[col].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    stats = {
        'nb_cartes'     : len(cartes),
        'nb_lignes'     : sum(len(c['lignes']) for c in cartes.values()),
        'total_debit'   : sum(c['total'] for c in cartes.values()),
        'alertes'       : alertes,
        'cartes'        : [f"{c['titulaire']} ({num})" for num, c in cartes.items()],
        'nom_fichier'   : f"SOGEPA_CB_{__import__('calendar').month_abbr[mois].upper()}_{annee}.xlsx",
    }
    return buf, stats
