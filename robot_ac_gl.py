import pandas as pd
import re
import shutil
from difflib import SequenceMatcher
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from io import BytesIO

MOIS_NOMS = {
    1:'Janvier', 2:'Février', 3:'Mars', 4:'Avril', 5:'Mai', 6:'Juin',
    7:'Juillet', 8:'Août', 9:'Septembre', 10:'Octobre', 11:'Novembre', 12:'Décembre'
}
MOIS_COL = {1:7,2:8,3:9,4:10,5:11,6:12,7:13,8:14,9:15,10:16,11:17,12:18}

KW = [
    (r'GAPAYLO',                                         '46710000', 'GAPAYLO',     'GAPAYLO'),
    (r'VIR M\.A CIC|VIRT M\.A|MICHEL AZRAN|VIR MA CIC', '45510000', 'MA',          'M. Michel Azran'),
    (r'COMMISSION.{0,15}INTERVENTION',                   '62700002', 'BANQUE',      'Commission CIC'),
    (r'FACT SGT|FACTURE CIC',                            '62700001', 'BANQUE',      'Facture CIC'),
    (r'FRAIS BANCAIRES',                                 '62700000', 'BANQUE',      'Frais bancaires'),
    (r'PRLV SEPA GENERALI',                              '61610003', 'ASSURANCE',   'Generali'),
    (r'PRLV SEPA.*AXA ',                                 '61610007', 'ASSURANCE',   'AXA'),
    (r'PRLV SEPA SWISSLIFE',                             '401SWISS', 'ASSURANCE',   'Swisslife'),
    (r'PRLV SEPA EDF',                                   '401EDF',   'FOURNISSEUR', 'EDF'),
    (r'PRLV SEPA UGIP',                                  '66110101', 'ASSURANCE',   'UGIP'),
    (r'IMMOGIM BAGNEUX|IMMOGIM B',                       '401IMMOB', 'FOURNISSEUR', 'Immogim Bagneux'),
    (r'PLV PARTIEL ECH|ECH PRET CAP|RB\.IMP',            None,       'PRET',        'Prêt — robot LT'),
    (r'DIRECTION GENERALE',                               None,       'IMPOT',       'Impôt — ignoré'),
    (r'VIR CAF DU VAL|VIR CAF SEINE',                   'CAF',       'CAF',         'Virement CAF'),
]
CORR = {
    'MRABAT':      ('411MRABA', 'MRABA B. & ZOUAOUI M.'),
    'ZOUAOUI':     ('411MRABA', 'MRABA B. & ZOUAOUI M.'),
    'SPETEBROODT': ('411VUONG', 'VUONG Z. & SPETEBROODT P.'),
    'NAVARRETE':   ('411NAVAE', 'NAVARETTE ERVIN'),
    'NAVARETTE':   ('411NAVAE', 'NAVARETTE ERVIN'),
}
NOUV = {
    'EXPERTS CHAUFFAGISTES': ('401EXPER', 'LES EXPERTS CHAUFFAGISTES'),
    'MT DESIGN':             ('401MTDES', 'MT DESIGN'),
}

def _kw(lib):
    for p,c,t,l in KW:
        if re.search(p, lib, re.IGNORECASE): return c,t,l
    return None,None,None

def _corr(lib):
    for m,(c,n) in CORR.items():
        if m.upper() in lib.upper(): return c,'LOYER',n
    return None,None,None

def _score(lib, nom):
    l=lib.upper(); n=nom.upper()
    s=SequenceMatcher(None,l,n).ratio()
    mots=[m for m in re.split(r'\W+',n) if len(m)>3]
    t=sum(1 for m in mots if m in l or l.find(m[:6])>=0)
    return max(s,(t/max(len(mots),1))*0.9)

def _match(lib, loc_p, four_p, ass_p):
    bs,bn,bc,bt=0,'','',''
    for p,(c,n) in NOUV.items():
        if p.upper() in lib.upper(): return c,'FOURNISSEUR',n,1.0
    for _,r in loc_p.iterrows():
        s=_score(lib,r['LIBELLE'])
        if s>bs: bs,bn,bc,bt=s,r['LIBELLE'],r['COMPTE'],'LOYER'
    for _,r in four_p.iterrows():
        s=_score(lib,r['LIBELLE'])
        if s>bs: bs,bn,bc,bt=s,r['LIBELLE'],r['COMPTE'],'FOURNISSEUR'
    for _,r in ass_p.iterrows():
        s=_score(lib,r['LIBELLE'])
        if s>bs: bs,bn,bc,bt=s,r['LIBELLE'],r['COMPTE'],'ASSURANCE'
    return bc,bt,bn,bs

def run(releve_io, pivot_io, loyers_io, mois, annee=2026, seuil=0.52):
    # Pivot
    df_p=pd.read_excel(pivot_io, header=None)
    df_p.columns=['COMPTE','LIBELLE','C','D']
    df_p=df_p[['COMPTE','LIBELLE']].dropna(subset=['LIBELLE'])
    df_p['COMPTE']=df_p['COMPTE'].astype(str).str.strip()
    df_p['LIBELLE']=df_p['LIBELLE'].astype(str).str.strip()
    loc_p =df_p[df_p['COMPTE'].str.startswith('411')].copy()
    four_p=df_p[df_p['COMPTE'].str.startswith('401')].copy()
    ass_p =df_p[df_p['COMPTE'].str.startswith('616')].copy()

    # Relevé
    df_r=pd.read_excel(releve_io, sheet_name=1, header=None)
    df_r.columns=df_r.iloc[4]
    df_r=df_r.iloc[5:].reset_index(drop=True)
    df_r.columns=['DATE','VALEUR','LIBELLE','DEBIT','CREDIT','SOLDE','DEV']
    df_r=df_r[df_r['LIBELLE'].notna()].copy()
    df_r['DATE']  =pd.to_datetime(df_r['DATE'],  errors='coerce')
    df_r['DEBIT'] =pd.to_numeric(df_r['DEBIT'],  errors='coerce').fillna(0)
    df_r['CREDIT']=pd.to_numeric(df_r['CREDIT'], errors='coerce').fillna(0)
    df_r=df_r[(df_r['DATE'].dt.month==mois)&(df_r['DATE'].dt.year==annee)].copy()

    if df_r.empty:
        return None, None, None, f"Aucune transaction pour {MOIS_NOMS[mois]} {annee}"

    # Traitement
    resultats=[]; alertes=[]
    for _,row in df_r.iterrows():
        lib=str(row['LIBELLE']).strip()
        date=row['DATE']
        montant=row['CREDIT'] if row['CREDIT']>0 else -row['DEBIT']
        sens='CREDIT' if row['CREDIT']>0 else 'DEBIT'

        c_kw,t_kw,l_kw=_kw(lib)
        if t_kw in('PRET','IMPOT'):
            resultats.append({'DATE':date,'LIBELLE':lib,'MONTANT':montant,'SENS':sens,'TYPE':t_kw,'COMPTE':'—','NOM':l_kw,'STATUT':'IGNORE'}); continue
        if c_kw:
            resultats.append({'DATE':date,'LIBELLE':lib,'MONTANT':montant,'SENS':sens,'TYPE':t_kw,'COMPTE':c_kw,'NOM':l_kw,'STATUT':'OK'}); continue
        c_c,t_c,n_c=_corr(lib)
        if c_c:
            resultats.append({'DATE':date,'LIBELLE':lib,'MONTANT':montant,'SENS':sens,'TYPE':t_c,'COMPTE':c_c,'NOM':n_c,'STATUT':'OK'}); continue
        mp=re.search(r'PRLVT\s+LOYER\s+([A-Z]+(?:\s+[A-Z]+)?)',lib,re.IGNORECASE)
        if mp:
            c_m,t_m,n_m,s_m=_match(mp.group(1).strip(),loc_p,four_p,ass_p)
            if s_m>=seuil:
                resultats.append({'DATE':date,'LIBELLE':lib,'MONTANT':montant,'SENS':sens,'TYPE':'LOYER','COMPTE':c_m,'NOM':n_m,'STATUT':'OK'}); continue
        c_m,t_m,n_m,s_m=_match(lib,loc_p,four_p,ass_p)
        if s_m>=seuil:
            resultats.append({'DATE':date,'LIBELLE':lib,'MONTANT':montant,'SENS':sens,'TYPE':t_m,'COMPTE':c_m,'NOM':n_m,'STATUT':'OK'})
        else:
            resultats.append({'DATE':date,'LIBELLE':lib,'MONTANT':montant,'SENS':sens,'TYPE':'?','COMPTE':'???','NOM':'NON IDENTIFIÉ','STATUT':'ALERTE'})
            alertes.append({'DATE':date,'LIBELLE':lib,'MONTANT':montant,'SENS':sens})

    df_res=pd.DataFrame(resultats)

    # Export 1 : Intégration logiciel
    rows=[]
    for _,r in df_res[df_res.STATUT=='OK'].iterrows():
        if r['COMPTE'] in('—','???','CAF'): continue
        d=r['DATE'].strftime('%Y-%m-%d') if pd.notna(r['DATE']) else ''
        m=abs(r['MONTANT']); lib=str(r['LIBELLE'])[:40]
        piece='PRLVT' if re.search(r'PRLV|PRLVT',lib,re.I) else ('VIRT' if 'VIR' in lib.upper() else 'AUTRE')
        d1=m if r['SENS']=='DEBIT' else 0; c1=m if r['SENS']=='CREDIT' else 0
        rows.append([d,'BQ0','47100000',piece,lib,round(d1,2),round(c1,2)])
        rows.append([d,'BQ0','51200000',piece,lib,round(c1,2),round(d1,2)])
    df_integ=pd.DataFrame(rows,columns=['DATE','CODE BANQUE','COMPTE','PIECE','LIBELLE','DEBIT','CREDIT'])
    buf1=BytesIO()
    with pd.ExcelWriter(buf1,engine='openpyxl') as w:
        df_integ.to_excel(w,index=False,sheet_name='Feuil1')
        ws=w.sheets['Feuil1']
        for c in ws[1]: c.fill=PatternFill('solid',fgColor='1E3A5F'); c.font=Font(bold=True,color='FFFFFF')
        f1=PatternFill('solid',fgColor='EEF2FF'); f2=PatternFill('solid',fgColor='FFFFFF')
        for i,row in enumerate(ws.iter_rows(min_row=2)):
            for c in row: c.fill=f1 if i%2==0 else f2
        for col,w2 in [('A',14),('B',12),('C',12),('D',10),('E',45),('F',12),('G',12)]:
            ws.column_dimensions[col].width=w2
    buf1.seek(0)

    # Export 2 : Tableau loyers
    buf2=BytesIO()
    loyers_io.seek(0)
    wb=load_workbook(loyers_io)
    sheet=next((s for s in wb.sheetnames if 'EMJJ' in s.upper() and 'MIRA' not in s.upper()), wb.sheetnames[0])
    ws_l=wb[sheet]
    COL_M=MOIS_COL[mois]
    loyers_ok=df_res[(df_res.STATUT=='OK')&(df_res.TYPE=='LOYER')].copy()
    fill_ok=PatternFill('solid',fgColor='C8E6C9'); fill_d=PatternFill('solid',fgColor='E8F5E9'); font_ok=Font(color='1B5E20',bold=True)
    loc_rows={}; cur_loc=None
    for row in ws_l.iter_rows():
        col2=ws_l.cell(row=row[0].row,column=2).value
        col4=ws_l.cell(row=row[0].row,column=4).value
        if col2 and 'Locataire' in str(col2):
            cur_loc=str(ws_l.cell(row=row[0].row,column=4).value or '').strip()
        if col4 and 'Règlement locataire' in str(col4) and cur_loc:
            if cur_loc not in loc_rows: loc_rows[cur_loc]=[]
            loc_rows[cur_loc].append(row[0].row)
    maj=[]
    for _,lr in loyers_ok.iterrows():
        nom_id=lr['NOM'].upper()
        montant=-(abs(lr['MONTANT']))
        date_s=lr['DATE'].strftime('%d/%m/%Y') if pd.notna(lr['DATE']) else ''
        mode='Virement' if 'VIR' in str(lr['LIBELLE']).upper() else 'Prélèvement'
        for nom_loc,row_list in loc_rows.items():
            mots=[m for m in re.split(r'\W+',nom_id) if len(m)>3]
            if any(m in nom_loc.upper() for m in mots):
                for row_idx in row_list:
                    cm=ws_l.cell(row=row_idx,column=COL_M)
                    if cm.value in(None,0,'','-'):
                        cm.value=round(montant,2); cm.fill=fill_ok; cm.font=font_ok
                        ws_l.cell(row=row_idx+1,column=COL_M).value=date_s
                        ws_l.cell(row=row_idx+1,column=COL_M).fill=fill_d
                        ws_l.cell(row=row_idx+2,column=COL_M).value=mode
                        ws_l.cell(row=row_idx+2,column=COL_M).fill=fill_d
                        maj.append(f"{nom_loc} → {montant:.2f}€")
                        break
                break
    wb.save(buf2); buf2.seek(0)

    # Export 3 : Alertes
    buf3=BytesIO()
    if alertes:
        df_a=pd.DataFrame(alertes)
        df_a['DATE']=df_a['DATE'].dt.strftime('%d/%m/%Y')
        df_a.columns=['DATE','LIBELLÉ','MONTANT','SENS']
        df_a['ACTION']='À saisir manuellement'
        with pd.ExcelWriter(buf3,engine='openpyxl') as w:
            df_a.to_excel(w,index=False,sheet_name='Alertes')
            ws_a=w.sheets['Alertes']
            for c in ws_a[1]: c.fill=PatternFill('solid',fgColor='C62828'); c.font=Font(bold=True,color='FFFFFF')
            for row in ws_a.iter_rows(min_row=2):
                for c in row: c.fill=PatternFill('solid',fgColor='FFF3E0')
            ws_a.column_dimensions['B'].width=65
    else:
        import openpyxl as ox
        wb_a=ox.Workbook(); ws_a=wb_a.active; ws_a.title='Alertes'
        ws_a.append(['RÉSULTAT'])
        ws_a.append(['✅ Aucune alerte — tout identifié automatiquement'])
        wb_a.save(buf3)
    buf3.seek(0)

    return df_res, (buf1,buf2,buf3), maj, None
